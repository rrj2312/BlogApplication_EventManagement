# events/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from django.db.models import Count
from django.utils import timezone
import json
import openpyxl
from .models import Event, Registration, Attendance
from .forms import RegistrationForm

def home(request):
    events = Event.objects.filter(is_active=True).order_by('-date')
    return render(request, 'events/home.html', {'events': events})

def event_detail(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    return render(request, 'events/event_detail.html', {'event': event})

def register(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            registration = form.save(commit=False)
            registration.event = event
            
            # Check if email already registered for this event
            if Registration.objects.filter(event=event, email=registration.email).exists():
                messages.error(request, 'This email is already registered for this event.')
                return render(request, 'events/register.html', {'form': form, 'event': event})
            
            registration.save()
            messages.success(request, f'Successfully registered! Your registration ID is: {registration.registration_id}')
            return redirect('registration_success', registration_id=registration.registration_id)
    else:
        form = RegistrationForm()
    
    return render(request, 'events/register.html', {'form': form, 'event': event})

def registration_success(request, registration_id):
    registration = get_object_or_404(Registration, registration_id=registration_id)
    return render(request, 'events/registration_success.html', {'registration': registration})

def qr_scanner(request):
    return render(request, 'events/qr_scanner.html')

@csrf_exempt
def scan_qr(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            qr_data = data.get('qr_data', '')
            
            # Parse QR data (format: registration_id|event_id)
            if '|' not in qr_data:
                return JsonResponse({'success': False, 'message': 'Invalid QR code format'})
            
            registration_id, event_id = qr_data.split('|')
            
            # Find registration
            try:
                registration = Registration.objects.get(
                    registration_id=registration_id,
                    event_id=event_id
                )
            except Registration.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Registration not found'})
            
            # Check if already attended
            if hasattr(registration, 'attendance'):
                return JsonResponse({
                    'success': False,
                    'message': f'Already marked attendance at {registration.attendance.timestamp.strftime("%Y-%m-%d %H:%M")}'
                })
            
            # Create attendance record
            attendance = Attendance.objects.create(
                registration=registration,
                scanned_by='QR Scanner'
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Attendance marked for {registration.name}',
                'participant': {
                    'name': registration.name,
                    'email': registration.email,
                    'registration_id': registration.registration_id,
                    'event': registration.event.name,
                    'timestamp': attendance.timestamp.strftime('%Y-%m-%d %H:%M:%S')
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Invalid JSON data'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

def admin_dashboard(request):
    events = Event.objects.annotate(
        total_registrations=Count('registrations'),
        total_attendance=Count('registrations__attendance')
    ).order_by('-date')
    
    recent_attendance = Attendance.objects.select_related('registration', 'registration__event').order_by('-timestamp')[:10]
    
    context = {
        'events': events,
        'recent_attendance': recent_attendance,
        'total_events': Event.objects.count(),
        'total_registrations': Registration.objects.count(),
        'total_attendance': Attendance.objects.count(),
    }
    
    return render(request, 'events/admin_dashboard.html', context)

def event_attendance_live(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    registrations = Registration.objects.filter(event=event).select_related('attendance').order_by('name')
    
    context = {
        'event': event,
        'registrations': registrations,
    }
    
    return render(request, 'events/event_attendance_live.html', context)

def export_attendance(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{event.name} Attendance"
    
    # Headers
    headers = ['Name', 'Email', 'Registration ID', 'Status', 'Timestamp']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    registrations = Registration.objects.filter(event=event).select_related('attendance')
    for row, registration in enumerate(registrations, 2):
        ws.cell(row=row, column=1, value=registration.name)
        ws.cell(row=row, column=2, value=registration.email)
        ws.cell(row=row, column=3, value=registration.registration_id)
        
        if hasattr(registration, 'attendance'):
            ws.cell(row=row, column=4, value='Present')
            ws.cell(row=row, column=5, value=registration.attendance.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        else:
            ws.cell(row=row, column=4, value='Absent')
            ws.cell(row=row, column=5, value='')
    
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{event.name}_attendance.xlsx"'
    
    wb.save(response)
    return response

# AJAX view for live updates
def get_live_attendance(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    registrations = Registration.objects.filter(event=event).select_related('attendance').order_by('name')
    
    data = {
        'total_registrations': registrations.count(),
        'total_attendance': registrations.filter(attendance__isnull=False).count(),
        'registrations': []
    }
    
    for reg in registrations:
        reg_data = {
            'name': reg.name,
            'email': reg.email,
            'registration_id': reg.registration_id,
            'is_attended': hasattr(reg, 'attendance'),
            'timestamp': reg.attendance.timestamp.strftime('%Y-%m-%d %H:%M:%S') if hasattr(reg, 'attendance') else None
        }
        data['registrations'].append(reg_data)
    
    return JsonResponse(data)